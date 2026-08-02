from __future__ import annotations

import csv
import html
import io
import json
import posixpath
import re
import sqlite3
import subprocess
import tempfile
import tomllib
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import Element

from quill.core.document import Document
from quill.core.epub import load_epub_book, render_epub_book
from quill.core.pdf_ocr_install import PACK_ASSISTED_SUFFIXES, extraction_pack_missing
from quill.core.safe_archive import check_zip_safety, open_zip
from quill.core.safe_xml import fromstring as safe_xml_fromstring
from quill.io.docx_text import read_docx_text, read_docx_xml_text
from quill.io.markitdown_bridge import convert_with_markitdown
from quill.io.pages import read_pages_document
from quill.io.pdf import extract_pdf_outline, extract_pdf_text, format_pdf_document
from quill.io.rtf import read_rtf_document
from quill.io.text import read_text_document

# PERF-11: spreadsheet reads are bounded so a huge sheet cannot exhaust memory.
# openpyxl's read-only mode streams rows lazily; these caps stop iteration early
# (and trim very wide rows), so an extract never materializes an unbounded sheet.
_SPREADSHEET_MAX_ROWS = 50
_SPREADSHEET_MAX_COLS = 20


def read_structured_document(
    path: Path,
    encoding: str = "utf-8",
    *,
    docx_engine: str = "auto",
    pdf_password: str | None = None,
) -> Document:
    suffix = path.suffix.lower()
    if suffix in {".sqlite", ".db"}:
        text = _format_sqlite(path)
        metadata = {"source_kind": "sqlite", "engine": "sqlite", "quality_score": 100}
    elif suffix in {".docx", ".docm"}:
        # .docm is a macro-enabled Word file: byte-for-byte the same OOXML
        # container as .docx, so the same reader chain handles it. QUILL reads
        # its text (it never executes macros); saving routes through Save As
        # like any other imported binary.
        _guard_office_zip(path)
        document = _read_docx(path, engine=docx_engine)
        if document is not None:
            if suffix == ".docm":
                document.source_metadata["source_kind"] = "docm"
            return document
        # #1279: python-docx is a base dependency, so the fallback reads headings,
        # lists, and tables rather than a flat paragraph dump; the hand-rolled XML
        # walk stays as the last resort for a document python-docx cannot open.
        rendered = read_docx_text(path)
        engine = "python-docx" if rendered is not None else "docx"
        text = rendered if rendered is not None else read_docx_xml_text(path)
        metadata = {"source_kind": suffix.lstrip("."), "engine": engine, "quality_score": 100}
    elif suffix == ".doc":
        document = _read_via_markitdown(path, "doc", fallback_engine="doc")
        if document is not None:
            return document
        text, metadata = _read_legacy_office_via_libreoffice(
            path, converted_suffix=".docx", source_kind="doc", fallback_engine="doc"
        )
    elif suffix == ".epub":
        text = render_epub_book(load_epub_book(path))
        metadata = {"source_kind": "epub", "engine": "epub", "quality_score": 100}
    elif suffix == ".pdf":
        text, metadata = _format_pdf(path, password=pdf_password)
    elif suffix == ".pages":
        doc = read_pages_document(path)
        text = doc.text
        metadata = doc.source_metadata
    elif suffix == ".odt":
        text = _format_odt(path)
        metadata = {"source_kind": "odt", "engine": "odt", "quality_score": 100}
    elif suffix == ".rtf":
        return read_rtf_document(path)
    elif suffix == ".pptx":
        _guard_office_zip(path)
        document = _read_via_markitdown(path, "pptx", fallback_engine="pptx")
        if document is not None:
            return document
        text = _format_pptx(path)
        metadata = {"source_kind": "pptx", "engine": "pptx", "quality_score": 100}
    elif suffix == ".ppt":
        document = _read_via_markitdown(path, "ppt", fallback_engine="ppt")
        if document is not None:
            return document
        text, metadata = _read_legacy_office_via_libreoffice(
            path, converted_suffix=".pptx", source_kind="ppt", fallback_engine="ppt"
        )
    elif suffix in {".xlsx", ".xls"}:
        _guard_office_zip(path)
        document = _read_spreadsheet_via_markitdown(path)
        if document is not None:
            return document
        text, metadata = _format_spreadsheet(path)
    else:
        # #1278: read_text_document strips a UTF-8 BOM (remembering utf-8-sig so a
        # save re-adds it), keeps the original line ending, and degrades to
        # cp1252/latin-1. The old path.read_text left the BOM in the text, so a
        # BOM-prefixed .json/.xml/.toml failed to parse (and crashed before #1228).
        loaded = read_text_document(path, encoding=encoding)
        text, metadata = _format_structured_text(loaded.text, suffix)
        return Document(
            text=text,
            path=path,
            modified=False,
            encoding=loaded.encoding,
            line_ending=loaded.line_ending,
            source_metadata=metadata,
        )

    # #1279: a fallback read still succeeds, but MarkItDown reads these formats
    # better -- flag it so the intake summary points at the one-click download.
    # Not for a Word file python-docx already read well: that path is good enough
    # that nagging about a 150 MB download would be noise, not help.
    if (
        suffix in PACK_ASSISTED_SUFFIXES
        and metadata.get("engine") != "python-docx"
        and extraction_pack_missing()
    ):
        metadata["extraction_pack_missing"] = True
    line_ending = "\r\n" if "\r\n" in text else "\n"
    return Document(
        text=text,
        path=path,
        modified=False,
        encoding=encoding,
        line_ending=line_ending,
        source_metadata=metadata,
    )


def _guard_office_zip(path: Path) -> None:
    """Reject a decompression bomb before a primary OOXML reader expands it.

    The office formats (.docx/.docm/.pptx/.xlsx) are ZIP containers, and the
    preferred readers (MarkItDown, openpyxl, python-docx) all expand entries
    without a size guard of their own. Running :func:`check_zip_safety` up front
    refuses a crafted archive whose declared expansion exceeds the safe limits
    before any entry is read, raising
    :class:`~quill.core.safe_archive.DecompressionBombError`. A non-ZIP file (a
    legacy .doc/.xls, or a corrupt upload) is left for the downstream reader to
    diagnose -- only a genuine bomb is refused here.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            check_zip_safety(archive)
    except (zipfile.BadZipFile, OSError):
        return


def _read_via_markitdown(
    path: Path,
    source_kind: str,
    *,
    fallback_engine: str,
    quality_score: int = 90,
) -> Document | None:
    try:
        text = convert_with_markitdown(path)
    except (ImportError, ValueError, RuntimeError):
        return None
    return Document(
        text=text,
        path=path,
        modified=False,
        encoding="utf-8",
        line_ending="\n",
        source_metadata={
            "source_kind": source_kind,
            "engine": "markitdown",
            "quality_score": quality_score,
            "fallback_engine": fallback_engine,
        },
    )


def _read_docx(path: Path, *, engine: str = "auto") -> Document | None:
    """Read a ``.docx`` via the preferred engine chain (``docx_read_engine``).

    ``auto`` and ``markitdown`` try MarkItDown first. ``pandoc`` converts to
    Markdown with Pandoc (richer structure: footnotes and complex tables
    survive better) and degrades to the MarkItDown chain when Pandoc is
    unavailable or fails, so an engine preference never fails an open.
    Returning ``None`` lets the caller fall back to the raw paragraph extract.
    """
    if engine == "pandoc":
        try:
            from quill.io import pandoc as pandoc_mod

            result = pandoc_mod.convert_document_with_pandoc(path, "markdown", from_format="docx")
            text = result.text
            if text.strip():
                return Document(
                    text=text.rstrip() + "\n",
                    path=path,
                    modified=False,
                    encoding="utf-8",
                    line_ending="\n",
                    source_metadata={
                        "source_kind": "docx",
                        "engine": "pandoc",
                        "quality_score": 90,
                        "fallback_engine": "docx",
                    },
                )
        except Exception:  # noqa: BLE001 - degrade to the default chain
            pass
    return _read_via_markitdown(path, "docx", fallback_engine="docx")


def _read_spreadsheet_via_markitdown(path: Path) -> Document | None:
    document = _read_via_markitdown(
        path,
        path.suffix.lower().lstrip("."),
        fallback_engine="csv table",
    )
    if document is not None:
        return document
    if path.suffix.lower() == ".xls":
        return _read_spreadsheet_via_libreoffice(path)
    return None


def _read_spreadsheet_via_libreoffice(path: Path) -> Document | None:
    from quill.core.external_tools import libreoffice_executable

    soffice = libreoffice_executable()
    if soffice is None:
        return None
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            converted_path = tmpdir_path / f"{path.stem}.xlsx"
            subprocess.run(
                [
                    soffice,
                    "--headless",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(tmpdir_path),
                    str(path),
                ],
                check=True,
                capture_output=True,
                timeout=45,
            )
            if not converted_path.exists():
                return None
            _guard_office_zip(converted_path)
            try:
                from openpyxl import load_workbook  # type: ignore[import-not-found]
            except ImportError:
                return None
            return _read_workbook_as_text(path, load_workbook(str(converted_path), data_only=True))
    except (FileNotFoundError, subprocess.SubprocessError, OSError):
        return None


def _read_legacy_office_via_libreoffice(
    path: Path,
    *,
    converted_suffix: str,
    source_kind: str,
    fallback_engine: str,
) -> tuple[str, dict[str, object]]:
    from quill.core.external_tools import libreoffice_executable

    soffice = libreoffice_executable()
    if soffice is None:
        return _missing_legacy_office_text(path, source_kind)
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            converted_path = tmpdir_path / f"{path.stem}{converted_suffix}"
            subprocess.run(
                [
                    soffice,
                    "--headless",
                    "--convert-to",
                    converted_suffix.lstrip("."),
                    "--outdir",
                    str(tmpdir_path),
                    str(path),
                ],
                check=True,
                capture_output=True,
                timeout=45,
            )
            if not converted_path.exists():
                return _missing_legacy_office_text(path, source_kind)
            text = convert_with_markitdown(converted_path)
            return text, {
                "source_kind": source_kind,
                "engine": "markitdown",
                "quality_score": 85,
                "fallback_engine": fallback_engine,
                "conversion_engine": "libreoffice",
            }
    except (
        FileNotFoundError,
        subprocess.SubprocessError,
        OSError,
        ImportError,
        ValueError,
        RuntimeError,
    ):
        return _missing_legacy_office_text(path, source_kind)


def _format_spreadsheet(path: Path) -> tuple[str, dict[str, object]]:
    # Detect ZIP-level corruption before attempting openpyxl so the "corrupted"
    # message reaches the user regardless of whether openpyxl is installed, and
    # refuse a decompression bomb (check_zip_safety) before openpyxl expands it.
    try:
        with zipfile.ZipFile(path) as archive:
            check_zip_safety(archive)
    except zipfile.BadZipFile:
        return _corrupt_spreadsheet_text(path)

    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:
        return _missing_spreadsheet_text(path)

    try:
        workbook = load_workbook(str(path), data_only=True, read_only=True)
    except zipfile.BadZipFile:
        return _corrupt_spreadsheet_text(path)
    except Exception:
        return _missing_spreadsheet_text(path)

    try:
        text = _read_workbook_as_text(path, workbook).text
    finally:
        workbook.close()
    return text, {
        "source_kind": path.suffix.lower().lstrip("."),
        "engine": "openpyxl",
        "quality_score": 90,
    }


def _read_workbook_as_text(path: Path, workbook: Any) -> Document:
    # #1279: no "# Spreadsheet Extract: <name>" banner. The per-sheet headings stay:
    # they are structure the workbook has, not text about the extract.
    lines: list[str] = []
    for worksheet in workbook.worksheets:
        lines.append(f"## Sheet: {worksheet.title}")
        rows = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            rows.append(["" if cell is None else str(cell) for cell in row[:_SPREADSHEET_MAX_COLS]])
            if row_index >= _SPREADSHEET_MAX_ROWS:
                break
        lines.extend(_rows_to_markdown_table(rows) or ["(no extractable cells)"])
        lines.append("")
    text = "\n".join(lines).rstrip() + "\n"
    return Document(
        text=text,
        path=path,
        modified=False,
        encoding="utf-8",
        line_ending="\n",
        source_metadata={
            "source_kind": path.suffix.lower().lstrip("."),
            "engine": "openpyxl",
            "quality_score": 90,
        },
    )


def _rows_to_markdown_table(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []
    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    lines = ["| " + " | ".join(padded[0]) + " |"]
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for row in padded[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return lines


def _corrupt_spreadsheet_text(path: Path) -> tuple[str, dict[str, object]]:
    return (
        (
            f"({path.name} appears to be corrupted or is not a valid XLSX file.)\n\n"
            "Try opening and re-saving it in Excel or LibreOffice Calc before importing.\n"
        ),
        {
            "source_kind": path.suffix.lower().lstrip("."),
            "engine": "unavailable",
            "quality_score": 0,
        },
    )


def _missing_spreadsheet_text(path: Path) -> tuple[str, dict[str, object]]:
    return (
        (
            f"(Spreadsheet import not available for {path.name}.)\n\n"
            'Open Help > Download Optional Components and download "PDF and '
            'Office text extraction" for spreadsheet extraction.\n'
        ),
        {
            "source_kind": path.suffix.lower().lstrip("."),
            "engine": "unavailable",
            "quality_score": 0,
        },
    )


def _missing_legacy_office_text(path: Path, source_kind: str) -> tuple[str, dict[str, object]]:
    return (
        (
            f"({source_kind.upper()} import not available for {path.name}.)\n\n"
            'Open Help > Download Optional Components and download "PDF and '
            'Office text extraction", and install LibreOffice, for legacy '
            "Office extraction.\n"
        ),
        {
            "source_kind": source_kind,
            "engine": "unavailable",
            "quality_score": 0,
        },
    )


def _format_pdf(path: Path, *, password: str | None = None) -> tuple[str, dict[str, object]]:
    result = extract_pdf_text(path, password=password)
    text = format_pdf_document(result)
    metadata: dict[str, object] = {
        "source_kind": "pdf",
        "engine": result.engine,
        "quality_score": result.quality_score,
        "page_count": result.page_count,
        "extracted_pages": result.extracted_pages,
        "page_scores": result.page_scores,
    }
    # Import the PDF's embedded outline (Adobe bookmarks) as (title, page) pairs so
    # the open flow can turn them into QUILL bookmarks. Only when a reader actually
    # saw pages (page_count > 0) -- a locked/absent PDF has none. Best-effort.
    if result.page_count > 0:
        outline = extract_pdf_outline(path, password=password)
        if outline:
            metadata["pdf_outline"] = outline
    if result.quality_score >= 50 and result.text.strip():
        return text, metadata

    # A locked PDF has no text for MarkItDown to read either (it cannot supply the
    # password), and running it would only bury the "encrypted" signal the open
    # flow needs to prompt for a password. Return the encrypted result unchanged.
    if result.engine == "encrypted":
        return text, metadata

    markitdown_document = _read_via_markitdown(
        path, "pdf", fallback_engine=result.engine, quality_score=85
    )
    if markitdown_document is None:
        return text, metadata

    metadata.update({
        "engine": f"{result.engine} + markitdown",
        "quality_score": max(result.quality_score, 85),
        "markitdown_used": True,
    })
    return markitdown_document.text, metadata


def _format_structured_text(raw_text: str, suffix: str) -> tuple[str, dict[str, object]]:
    """Pretty-print / validate a structured text file, falling back to raw on failure.

    A malformed structured file (invalid JSON, TOML, XML, CSV, or notebook) must never
    crash the open flow — the user asked to *see* the file, most often precisely because
    it is broken and they want to read or repair it. When the format-specific parser
    raises, we return the file's raw text unchanged with a ``structured_parse_error``
    note in the metadata instead of letting the exception propagate (#1228).
    """
    formatters: dict[str, tuple[str, str, object]] = {
        ".json": ("json", "json", lambda t: _format_json(t)),
        ".toml": ("toml", "toml", lambda t: _validate_toml(t)),
        ".xml": ("xml", "xml", lambda t: _format_xml(t)),
        ".csv": ("csv", "delimited", lambda t: _format_delimited(t, delimiter=",")),
        ".tsv": ("tsv", "delimited", lambda t: _format_delimited(t, delimiter="\t")),
        ".ipynb": ("ipynb", "notebook", lambda t: _format_notebook(t)),
    }
    entry = formatters.get(suffix)
    if entry is None:
        return raw_text, {"source_kind": "text", "engine": "plain text", "quality_score": 100}
    source_kind, engine, formatter = entry
    try:
        text = formatter(raw_text)  # type: ignore[operator]
    except (ValueError, SyntaxError, csv.Error) as exc:
        # ValueError covers json.JSONDecodeError and tomllib.TOMLDecodeError;
        # SyntaxError covers ElementTree's ParseError; csv.Error covers delimited parsing.
        return (
            raw_text,
            {
                "source_kind": "text",
                "engine": "plain text",
                "quality_score": 0,
                "structured_parse_error": f"{type(exc).__name__}: {exc}",
            },
        )
    return text, {"source_kind": source_kind, "engine": engine, "quality_score": 100}


def _format_json(text: str) -> str:
    payload = json.loads(text)
    return json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n"


def _validate_toml(text: str) -> str:
    tomllib.loads(text)
    return text


def _format_xml(text: str) -> str:
    root = safe_xml_fromstring(text)
    ET.indent(root, space="  ")
    return ET.tostring(root, encoding="unicode") + "\n"


def _format_delimited(text: str, delimiter: str) -> str:
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, lineterminator="\n")
    for row in reader:
        writer.writerow(row)
    return output.getvalue()


def _format_notebook(text: str) -> str:
    payload = json.loads(text)
    cells = payload.get("cells", [])
    if not isinstance(cells, list):
        return "# Notebook\n\n(no cells)\n"
    lines = ["# Notebook", ""]
    for index, cell in enumerate(cells, start=1):
        if not isinstance(cell, dict):
            continue
        kind = str(cell.get("cell_type", "unknown"))
        source = cell.get("source", [])
        if isinstance(source, list):
            content = "".join(str(item) for item in source).strip()
        else:
            content = str(source).strip()
        lines.append(f"## Cell {index} ({kind})")
        lines.append(content or "(empty)")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _format_sqlite(path: Path) -> str:
    lines = [f"# SQLite Database: {path.name}", ""]
    with sqlite3.connect(path) as connection:
        cursor = connection.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type='table' AND name NOT LIKE 'sqlite_%' "
            "ORDER BY name"
        )
        table_names = [row[0] for row in cursor.fetchall()]
        if not table_names:
            lines.append("(no user tables)")
            return "\n".join(lines) + "\n"
        for table in table_names:
            cursor.execute(f'SELECT COUNT(*) FROM "{table}"')
            count = cursor.fetchone()[0]
            lines.append(f"- {table}: {count} row(s)")
    lines.append("")
    return "\n".join(lines)


def _format_odt(path: Path) -> str:
    with open_zip(path) as archive:
        content = archive.read("content.xml").decode("utf-8", errors="ignore")
    text = _strip_markup(content)
    # #1279: no synthetic banner -- see _format_docx.
    return (text or "(no extractable text)").rstrip() + "\n"


def _strip_markup(text: str) -> str:
    cleaned = re.sub(r"<[^>]+>", " ", text)
    cleaned = html.unescape(cleaned)
    return " ".join(cleaned.split())


_PPTX_NAMESPACES = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _format_pptx(path: Path) -> str:
    with open_zip(path) as archive:
        slide_paths = sorted(
            [
                name
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            ],
            key=_pptx_slide_sort_key,
        )
        if not slide_paths:
            return "(no extractable slides)\n"
        slide_rel_map = _load_pptx_slide_rel_map(archive)
        # #1279: no "# PPTX Extract: <name>" banner; the per-slide headings below
        # are the deck's own structure and stay.
        lines: list[str] = []
        for index, slide_path in enumerate(slide_paths, start=1):
            root = safe_xml_fromstring(archive.read(slide_path).decode("utf-8", errors="ignore"))
            title = _pptx_slide_title(root) or f"Slide {index}"
            lines.append(f"# {title}")
            body_lines = _pptx_slide_body_lines(root)
            lines.extend(body_lines or ["(no body text)"])

            table_blocks = _pptx_slide_table_blocks(root)
            if table_blocks:
                lines.append("")
                lines.append("## Tables")
                for table in table_blocks:
                    lines.extend(table)
                    lines.append("")

            notes_lines = _pptx_slide_notes_lines(archive, slide_path, slide_rel_map)
            if notes_lines:
                lines.append("")
                lines.append("## Notes")
                lines.extend(notes_lines)
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _pptx_slide_sort_key(path: str) -> tuple[int, str]:
    match = re.search(r"slide(\d+)\.xml$", path)
    if match is None:
        return (0, path)
    return (int(match.group(1)), path)


def _load_pptx_slide_rel_map(archive: zipfile.ZipFile) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for rel_path in [
        name
        for name in archive.namelist()
        if name.startswith("ppt/slides/_rels/slide") and name.endswith(".xml.rels")
    ]:
        try:
            root = safe_xml_fromstring(archive.read(rel_path).decode("utf-8", errors="ignore"))
        except ET.ParseError:
            continue
        slide_name = rel_path.replace("ppt/slides/_rels/", "").replace(".rels", "")
        for relation in root.findall("pr:Relationship", _PPTX_NAMESPACES):
            rel_type = relation.attrib.get("Type", "")
            target = relation.attrib.get("Target", "")
            if not rel_type.endswith("/notesSlide") or not target:
                continue
            source_part = f"ppt/slides/{slide_name}"
            normalized = _normalize_pptx_target(source_part, target)
            mapping[slide_name] = normalized
            break
    return mapping


def _normalize_pptx_target(source_part_path: str, target: str) -> str:
    base_dir = Path(source_part_path).parent.as_posix()
    return posixpath.normpath(posixpath.join(base_dir, target))


def _pptx_slide_title(root: Element) -> str:
    for shape in root.findall(".//p:sp", _PPTX_NAMESPACES):
        if not _pptx_shape_is_title(shape):
            continue
        paragraphs = _pptx_shape_paragraphs(shape)
        for paragraph in paragraphs:
            if paragraph.strip():
                return paragraph.strip()
    return ""


def _pptx_shape_is_title(shape: Element) -> bool:
    placeholder = shape.find(".//p:nvSpPr/p:nvPr/p:ph", _PPTX_NAMESPACES)
    if placeholder is None:
        return False
    placeholder_type = (placeholder.attrib.get("type") or "").strip()
    return placeholder_type in {"title", "ctrTitle", "subTitle"}


def _pptx_shape_paragraphs(shape: Element) -> list[str]:
    paragraphs: list[str] = []
    for paragraph in shape.findall(".//a:p", _PPTX_NAMESPACES):
        text = "".join(
            node.text or "" for node in paragraph.findall(".//a:t", _PPTX_NAMESPACES)
        ).strip()
        if text:
            paragraphs.append(text)
    return paragraphs


def _pptx_slide_body_lines(root: Element) -> list[str]:
    lines: list[str] = []
    for shape in root.findall(".//p:sp", _PPTX_NAMESPACES):
        if _pptx_shape_is_title(shape):
            continue
        for paragraph in shape.findall(".//a:p", _PPTX_NAMESPACES):
            text = "".join(
                node.text or "" for node in paragraph.findall(".//a:t", _PPTX_NAMESPACES)
            ).strip()
            if not text:
                continue
            level = 0
            paragraph_props = paragraph.find("a:pPr", _PPTX_NAMESPACES)
            if paragraph_props is not None:
                raw_level = paragraph_props.attrib.get("lvl")
                if raw_level and raw_level.isdigit():
                    level = int(raw_level)
            lines.append(f"{'  ' * max(0, level)}- {text}")
    return lines


def _pptx_slide_table_blocks(root: Element) -> list[list[str]]:
    tables: list[list[str]] = []
    for table in root.findall(".//a:tbl", _PPTX_NAMESPACES):
        rows: list[list[str]] = []
        for row in table.findall("a:tr", _PPTX_NAMESPACES):
            cells: list[str] = []
            for cell in row.findall("a:tc", _PPTX_NAMESPACES):
                text = "".join(
                    node.text or "" for node in cell.findall(".//a:t", _PPTX_NAMESPACES)
                ).strip()
                cells.append(text or " ")
            if cells:
                rows.append(cells)
        if not rows:
            continue
        columns = max(len(row) for row in rows)
        normalized = [row + [" "] * (columns - len(row)) for row in rows]
        header = normalized[0]
        block = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join("---" for _ in header) + " |",
        ]
        for body_row in normalized[1:]:
            block.append("| " + " | ".join(body_row) + " |")
        tables.append(block)
    return tables


def _pptx_slide_notes_lines(
    archive: zipfile.ZipFile,
    slide_path: str,
    slide_rel_map: dict[str, str],
) -> list[str]:
    slide_name = Path(slide_path).name
    notes_path = slide_rel_map.get(slide_name)
    if notes_path is None:
        return []
    if notes_path not in archive.namelist():
        return []
    try:
        root = safe_xml_fromstring(archive.read(notes_path).decode("utf-8", errors="ignore"))
    except ET.ParseError:
        return []
    lines: list[str] = []
    for paragraph in root.findall(".//a:p", _PPTX_NAMESPACES):
        text = "".join(
            node.text or "" for node in paragraph.findall(".//a:t", _PPTX_NAMESPACES)
        ).strip()
        if text:
            lines.append(text)
    return lines
